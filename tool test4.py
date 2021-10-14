# importing openpyxl module 
import openpyxl as xl; 
import re
# opening the source excel file

print("Select the Physical File : ")

fileName2 = f.askopenfilename()
#print(fileName2)

aa = fileName2.replace("/", "\\\\")
#print(aa)


# opening the source excel file 
filename =aa
wb1 = xl.load_workbook(filename) 
ws1 = wb1.worksheets[5] 




print("Select the Excel sheet where the auth model is to be generated : ")
fileName3 = f.askopenfilename()
#print(fileName3)

bb = fileName3.replace("/", "\\\\")
#print(bb)


# opening the destination excel file





filename1 =bb
wb2 = xl.load_workbook(filename1) 
ws2 = wb2.worksheets[5] 



# calculate total number of rows and 
# columns in source excel file

mr = ws1.max_row 
mc = ws1.max_column 





# copying the cell values from source 
# excel file to destination excel file
print("Enter the row number from where Physical Field Name copied to short name : ")
valrownum=int(input())

#Physical Field Name (E) to short name (B)
for i in range (2,mr+1):
        c = ws1.cell(row = i, column = valrownum)
        a1=c.value
        a2=str(a1)
        if a2!="None":
                stri=a2.lower()
                b=stri.replace(" ","")
                ws2.cell(row = i, column = 2).value = b
                ws2.cell(row = i, column = 6).value = c.value
        


print("Enter the row number from where Physical Display Name  copied to Display Name: ")
valrownum=int(input())


#Physical Display Name (F) to Display Name (F)
for i in range (2,mr+1):
        c = ws1.cell(row = i, column = valrownum)
        ws2.cell(row = i, column = 6).value = c.value









print("Enter the row number from where Physical / Technical Definition  copied to Help Text: ")
valrownum=int(input())

#Physical / Technical Definition (H) to Help Text(Q)
for i in range (2,mr+1):
        c = ws1.cell(row = i, column = valrownum)
        ws2.cell(row = i, column = 17).value = c.value

print("Enter the row number from where Datatype  copied to Datatype: ")
valrownum1=int(input())

#Datatype (M) to Datatype (D)
for i in range (2,mr+1):
        c = ws1.cell(row = i, column = valrownum1)
        a1=c.value
        a2=str(a1)
        if a2!="None":
                a3=a2.lower()
                if a3=="number":
                        ws2.cell(row = i, column = 4).value = "integer"


                elif a3=="char":
                        ws2.cell(row = i, column = 4).value = "string"
        
                elif a3.find('string')!=-1:
                        ws2.cell(row = i, column = 4).value = "string"

                else:
                        ws2.cell(row = i, column = 4).value = a3
                
print("Enter the row number from where Field Group  copied to GROUP: ")
valrownum=int(input())

#Field Group (U) to Group (G)
for i in range (2,mr+1):
        c = ws1.cell(row = i, column = valrownum)
        ws2.cell(row = i, column = 7).value = c.value

#print("Enter the row number from where Datatype  copied to Displaytype: ")
#valrownum1=int(input())


#Datatype (M) to DisplayType(H)
for i in range (2,mr+1):
        c = ws1.cell(row = i, column = valrownum1)
        a1=c.value
        a2=str(a1)
        if a2!="None":
                a3=a2.lower()
                if a3=="number":
                        ws2.cell(row = i, column = 8).value = "numerictextbox"


                elif a3=="char":
                        ws2.cell(row = i, column = 8).value = "textbox"
        
                elif a3.find('string')!=-1:
                        ws2.cell(row = i, column = 8).value = "textbox"

                elif a3=="string":
                        ws2.cell(row = i, column = 8).value = "textbox"

                else:
                        ws2.cell(row = i, column = 8).value = a3

print("Enter the row number from where Multi Language?  copied to Istranslatable: ")
valrownum=int(input())


#Multi Language? (Q) to Istranslatable 
for i in range (2,mr+1):
        c = ws1.cell(row = i, column = valrownum)
        ws2.cell(row = i, column = 10).value = c.value



#Is Mandatory (G) to Mandatory (k)
#for i in range (2,mr+1):
#        c = ws1.cell(row = i, column = 7)
#        ws2.cell(row = i, column = 11).value = c.value





        
        




       
# saving the destination excel file
wb2.save(str(filename1))




#------------ E-A-R sheet ------------------------------------------------------------------------------------------------------
print("Enter the row number for E-A-R sheet : ")

n1 = int(input());
n=n1-2


#Read P-I-A column to copy E-A-R entity and mapped attributes



filename =aa
wb1 = xl.load_workbook(filename) 
ws1 = wb1.worksheets[5]


mr = ws1.max_row 
mc = ws1.max_column



filename2=bb
wb3 = xl.load_workbook(filename2) 
ws3 = wb3.worksheets[9] 

mr1 = ws3.max_row 

z1="product"
z2="item"
z3="article"
z4="supplierarticle"



print("Enter the row number from where Physical Field Name  copied to MAPPED ATTRIBUTE and PIA column copied to Entity: ")
valrownum=int(input())
valrownum2=int(input())

for i in range (2,mr+1):
    c = ws1.cell(row = i, column = valrownum)  #attribute name column
    d = ws1.cell(row = i, column = valrownum2) #pia column
    
                       

    #convert to short name
    a1=c.value
    a2=str(a1)
    stri=a2.lower()
    b=stri.replace(" ","")

    #check the columnm
    x=d.value
    x1=str(x)
    y1=x1.lower()
    y2=y1.replace(" ","")
    y3=y2.replace('-','')
    y4=y3.replace(',','')
    y=str(y4)
    k=["pi","ip"]
    l=["ia","ai"]
    m=["pia" , "pai" , "iap" , "ipa" , "iap"]

    
    print(b,y)

    if y=="p":
            ws3.cell(row=i+n,column=2).value=z1         #attribute short name
            ws3.cell(row=i+n,column=4).value=b         #pia columnpaste

    elif y == "i":
            ws3.cell(row = i + n, column = 2).value = z2# attribute shortname
            ws3.cell(row = i + n, column = 4).value = b# pia column paste

    elif y == "a":
            ws3.cell(row = i + n, column = 2).value = z3# attribute shortname
            ws3.cell(row = i + n, column = 4).value = b# pia column paste
        
            ws3.cell(row = i + n+1, column = 2).value = z4# attribute shortname-----for supplierarticle
            ws3.cell(row = i + n+1, column = 4).value = b# pia column paste---------forsupplierarticle

            n=n+1


    elif y in k:
            ws3.cell(row = i + n, column = 4).value = b# pia column paste
            ws3.cell(row = i + n + 1, column = 4).value = b# pia column paste


            ws3.cell(row = i + n, column = 2).value = z1# attribute short name
            ws3.cell(row = i + n + 1, column = 2).value = z2# attribute shortname

            n = n + 1

    elif y in l:
            ws3.cell(row = i + n, column = 4).value = b# pia column paste
            ws3.cell(row = i + n + 1, column = 4).value = b# pia column paste
            ws3.cell(row = i + n + 2, column = 4).value = b# pia column paste


            ws3.cell(row = i + n, column = 2).value = z2# attribute shortname
            ws3.cell(row = i + n + 1, column = 2).value = z3# attribute shortname
            ws3.cell(row = i + n + 2, column = 2).value = z4# attribute shortname


            n = n + 2

    elif y in m:
            ws3.cell(row = i + n, column = 4).value = b# pia columnpaste
            ws3.cell(row = i + n, column = 2).value = z1# attribute shortname

            ws3.cell(row = i + n + 1, column = 4).value = b# pia column paste
            ws3.cell(row = i + n + 1, column = 2).value = z2# attribute shortname

            ws3.cell(row = i + n + 2, column = 4).value = b# pia column paste
            ws3.cell(row = i + n + 2, column = 2).value = z3# attribute shortname

            ws3.cell(row = i + n + 3, column = 4).value = b# pia column paste
            ws3.cell(row = i + n + 3, column = 2).value = z4# attribute shortname



            
            n=n+3
        







        
wb3.save(str(filename2))


